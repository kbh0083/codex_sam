from __future__ import annotations

import contextlib
import logging
import re
import struct
import threading
import warnings
from datetime import datetime
from pathlib import Path
import shutil
import xml.etree.ElementTree as ET

import openpyxl

from app.amount_normalization import normalize_excel_numeric_cell

try:
    import xlrd
except ImportError:  # pragma: no cover - optional dependency
    xlrd = None

logger = logging.getLogger(__name__)
_XLRD_SST_PATCH_LOCK = threading.RLock()


class ExcelDocumentLoaderMixin:
    """Excel 문서를 공통 raw_text로 변환하는 보조 로직.

    Excel은 PDF보다 구조가 명확한 편이지만, 시트가 여러 개이거나
    머리글/바닥글, 날짜 직렬값, 줄바꿈 셀 같은 요소가 섞이면
    그대로 문자열화했을 때 표 구조가 쉽게 깨질 수 있다.

    이 mixin의 목적은:
    - 시트 단위를 보존하고
    - 셀 값을 사람이 읽을 수 있는 문자열로 정리한 뒤
    - 후단 `DocumentLoader.build_markdown()`이 다시 해석할 수 있는
      `A | B | C` 형태의 raw_text로 맞추는 것이다.
    """

    def _load_workbook(self, file_path: Path) -> str:
        """`.xlsx/.xlsm` 계열 워크북을 읽어 시트별 raw_text를 만든다.

        `data_only=True`를 주는 이유는 수식 문자열이 아니라 계산된 값을 읽기 위해서다.
        추출 서비스 관점에서 필요한 것은 "표현식"이 아니라 "실제 숫자"이기 때문이다.
        """
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Cannot parse header or footer so it will be ignored",
                category=UserWarning,
            )
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        sections: list[str] = []
        logger.info("Opened workbook: sheets=%s", len(workbook.worksheets))
        for sheet in workbook.worksheets:
            rows: list[str] = [f"[SHEET {sheet.title}]"]
            for row in sheet.iter_rows():
                formatted = [self._format_openpyxl_cell(cell) for cell in row]
                if any(cell for cell in formatted):
                    rows.append(" | ".join(formatted))
            if len(rows) > 1:
                sections.append("\n".join(rows))
        logger.info("Extracted text from %s sheet section(s)", len(sections))
        if not sections:
            raise ValueError("No extractable text found in workbook.")
        return self.RAW_SECTION_DELIMITER.join(sections)

    def _load_legacy_workbook(self, file_path: Path) -> str:
        """`.xls`(BIFF) 워크북을 우선 로컬 기본 경로로 읽고, 실패 시 parser-style fallback을 탄다."""
        if xlrd is None:
            logger.info(
                "Legacy xls direct BIFF parser unavailable; trying parser-style fallback: file=%s",
                file_path,
            )
            return self._load_legacy_workbook_via_parser_fallback(file_path)

        try:
            workbook = xlrd.open_workbook(str(file_path), formatting_info=True)
        except Exception as exc:
            logger.info(
                "Legacy xls direct BIFF parser failed; trying parser-style fallback: file=%s error=%s",
                file_path,
                exc,
            )
            return self._load_legacy_workbook_via_parser_fallback(file_path)

        return self._render_legacy_workbook(
            workbook,
            include_hidden_state=True,
            log_prefix="xlrd BIFF parser",
        )

    @staticmethod
    def _format_xls_cell(sheet, row_idx: int, col_idx: int, datemode: int) -> str:
        """WAS parser와 같은 규칙으로 xlrd 셀 값을 문자열로 변환한다."""
        cell_type = sheet.cell_type(row_idx, col_idx)
        cell_value = sheet.cell_value(row_idx, col_idx)

        if xlrd is not None and cell_type == xlrd.XL_CELL_DATE:
            try:
                dt_tuple = xlrd.xldate_as_tuple(cell_value, datemode)
                year, month, day, hour, minute, second = dt_tuple
                if hour == 0 and minute == 0 and second == 0:
                    return f"{year:04d}-{month:02d}-{day:02d}"
                return (
                    f"{year:04d}-{month:02d}-{day:02d} "
                    f"{hour:02d}:{minute:02d}:{second:02d}"
                )
            except xlrd.XLDateError:
                return ExcelDocumentLoaderMixin._normalize_cell_text(str(cell_value))

        if xlrd is not None and cell_type == xlrd.XL_CELL_EMPTY:
            return ""

        if isinstance(cell_value, float) and cell_value.is_integer():
            return str(int(cell_value))

        return ExcelDocumentLoaderMixin._normalize_cell_text(str(cell_value))

    def _load_xls_as_xlsx(self, file_path: Path) -> str:
        """실제 xlsx인데 `.xls` 확장자인 파일을 xlsx 경로로 다시 읽는다."""
        tmp_path = file_path.with_suffix(".xlsx")
        try:
            shutil.copy2(file_path, tmp_path)
            return self._load_workbook(tmp_path)
        finally:
            tmp_path.unlink(missing_ok=True)

    def _load_legacy_workbook_via_parser_fallback(self, file_path: Path) -> str:
        """WAS parser `_extract_xls`와 같은 순서로 `.xls`를 열고, 결과를 markdown 친화적으로 정규화한다."""
        if self._is_xml_spreadsheet(file_path):
            return self._load_xml_spreadsheet(file_path)

        try:
            workbook = self._open_legacy_workbook_with_recovery(file_path)
        except Exception as exc:
            if self._is_xml_spreadsheet(file_path):
                return self._load_xml_spreadsheet(file_path)
            logger.info(
                "Parser-style xls fallback could not open BIFF workbook; trying xlsx path: file=%s error=%s",
                file_path,
                exc,
            )
            return self._load_xls_as_xlsx(file_path)

        return self._render_legacy_workbook(
            workbook,
            include_hidden_state=False,
            log_prefix="parser-style fallback",
        )

    def _open_legacy_workbook_with_recovery(self, file_path: Path):
        """parser-style BIFF open을 시도하고, SST underflow 계열 오류에서만 보수적 recovery를 적용한다."""
        try:
            return xlrd.open_workbook(str(file_path), formatting_info=False)
        except Exception as exc:
            if not self._is_biff_underflow_error(exc):
                raise

        file_bytes = file_path.read_bytes()
        padding = (-len(file_bytes)) % 512
        padded_bytes = file_bytes + (b"\x00" * padding)

        with self._patched_xlrd_sst_table():
            return xlrd.open_workbook(file_contents=padded_bytes, formatting_info=False)

    @staticmethod
    @contextlib.contextmanager
    def _patched_xlrd_sst_table():
        """`xlrd` SST 테이블 파싱 underflow를 빈 문자열 fallback으로 완화한다."""
        if xlrd is None:
            yield
            return

        unpack_sst_table = getattr(getattr(xlrd, "book", None), "unpack_SST_table", None)
        if unpack_sst_table is None:
            yield
            return

        def safe_unpack_sst_table(datatab, nstrings):
            try:
                return unpack_sst_table(datatab, nstrings)
            except struct.error as exc:
                if "unpack requires a buffer of 2 bytes" not in str(exc):
                    raise
                return ([""] * nstrings, {})

        with _XLRD_SST_PATCH_LOCK:
            original = xlrd.book.unpack_SST_table
            xlrd.book.unpack_SST_table = safe_unpack_sst_table
            try:
                yield
            finally:
                xlrd.book.unpack_SST_table = original

    @staticmethod
    def _is_biff_underflow_error(exc: Exception) -> bool:
        return isinstance(exc, struct.error) and "unpack requires a buffer of 2 bytes" in str(exc)

    def _render_legacy_workbook(self, workbook, *, include_hidden_state: bool, log_prefix: str) -> str:
        """워크북을 `[SHEET ...]` + pipe row raw_text로 정규화한다."""
        sections: list[str] = []
        extracted_sheet_names: list[str] = []
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            if include_hidden_state and getattr(sheet, "visibility", 0) != 0:
                continue

            extracted_sheet_names.append(sheet_name)
            rows: list[str] = [f"[SHEET {sheet_name}]"]
            hidden_cols = set()
            if include_hidden_state:
                hidden_cols = {
                    col_idx
                    for col_idx, col_info in getattr(sheet, "colinfo_map", {}).items()
                    if getattr(col_info, "hidden", 0)
                }

            for row_index in range(sheet.nrows):
                if include_hidden_state:
                    row_info = getattr(sheet, "rowinfo_map", {}).get(row_index)
                    if row_info and getattr(row_info, "hidden", 0):
                        continue

                formatted = [
                    self._format_xls_cell(sheet, row_index, col_index, workbook.datemode)
                    for col_index in range(sheet.ncols)
                    if col_index not in hidden_cols
                ]
                formatted = self._trim_trailing_empty_cells(formatted)
                if any(cell for cell in formatted):
                    rows.append(" | ".join(formatted))

            if len(rows) > 1:
                sections.append("\n".join(rows))

        logger.info(
            "Extracted text from legacy xls via %s: sheet_count=%s",
            log_prefix,
            len(extracted_sheet_names),
        )
        if not sections:
            raise ValueError("No extractable text found in legacy workbook.")
        return self.RAW_SECTION_DELIMITER.join(sections)

    @staticmethod
    def _is_xml_spreadsheet(file_path: Path) -> bool:
        """파일이 XML Spreadsheet 2003 형식인지 확인한다."""
        try:
            with file_path.open("rb") as handle:
                header = handle.read(512)
            return b"<?xml" in header and b"urn:schemas-microsoft-com:office:spreadsheet" in header
        except Exception:
            return False

    def _load_xml_spreadsheet(self, file_path: Path) -> str:
        """XML Spreadsheet 2003 형식에서 raw_text를 추출한다."""
        ns = {
            "ss": "urn:schemas-microsoft-com:office:spreadsheet",
        }
        tree = ET.parse(file_path)  # noqa: S314
        root = tree.getroot()

        sections: list[str] = []
        for worksheet in root.findall("ss:Worksheet", ns):
            sheet_name = worksheet.get(f"{{{ns['ss']}}}Name", "Sheet")
            rows: list[str] = [f"[SHEET {sheet_name}]"]

            table = worksheet.find("ss:Table", ns)
            if table is None:
                continue

            for row in table.findall("ss:Row", ns):
                cells = row.findall("ss:Cell", ns)
                values: list[str] = []
                for cell in cells:
                    data = cell.find("ss:Data", ns)
                    values.append(self._normalize_cell_text(data.text if data is not None and data.text else ""))
                values = self._trim_trailing_empty_cells(values)
                if any(value.strip() for value in values):
                    rows.append(" | ".join(values))

            if len(rows) > 1:
                sections.append("\n".join(rows))

        if not sections:
            raise ValueError("No extractable text found in legacy workbook.")
        return self.RAW_SECTION_DELIMITER.join(sections)

    @staticmethod
    def _format_cell(value: object) -> str:
        """openpyxl 셀 값을 사람이 읽을 수 있는 한 줄 문자열로 바꾼다."""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return ExcelDocumentLoaderMixin._normalize_cell_text(str(value))

    @staticmethod
    def _format_openpyxl_cell(cell) -> str:
        """openpyxl cell을 data_type-aware 문자열로 바꾼다."""
        value = cell.value
        if value is None:
            return ""
        if cell.is_date and isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if cell.data_type == "n" and not isinstance(value, bool):
            return ExcelDocumentLoaderMixin._normalize_cell_text(
                normalize_excel_numeric_cell(value)
            )
        return ExcelDocumentLoaderMixin._format_cell(value)

    @staticmethod
    def _normalize_cell_text(value: str) -> str:
        """엑셀 셀 내부 줄바꿈/여분 공백을 한 줄 텍스트로 정규화한다.

        다단 헤더 셀은 종종 `순유입금액\\n(확정)`처럼 개행을 포함한다.
        이 개행이 raw_text에 그대로 남으면 한 행이 여러 줄로 찢어져
        표 구조 추론과 coverage 계산이 크게 흔들릴 수 있다.
        """
        collapsed = re.sub(r"[\r\n\t]+", " ", value)
        collapsed = re.sub(r"\s{2,}", " ", collapsed)
        return collapsed.strip()

    @staticmethod
    def _trim_trailing_empty_cells(values: list[str]) -> list[str]:
        """행 끝의 빈 셀을 제거해 마지막 non-empty 셀에 pipe residue가 남지 않게 한다."""
        trimmed = list(values)
        while trimmed and not trimmed[-1]:
            trimmed.pop()
        return trimmed
